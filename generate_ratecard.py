import argparse
import copy
import datetime as dt
from pathlib import Path

import openpyxl


TARGET_SHEET = "5月刊例"
INACTIVE_SHEET = "下架留存"
OTHER_ORG_SHEET = "其他机构留存"
DUPLICATE_SHEET = "重复达人"
LOG_SHEET = "更新日志"

MANUAL_FIELDS = {
    "合作笔记案例参考",
    "返点",
    "备注",
    "机构",
    "负责人",
}


def normalize_header(value):
    if value is None:
        return ""
    return str(value).replace("\n", "").replace(" ", "").strip()


def normalize_key(value):
    if value is None:
        return ""
    text = str(value).strip()
    if text.endswith(".0"):
        text = text[:-2]
    return text


def parse_number(value):
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return value
    text = str(value).strip().replace(",", "")
    if not text or text in {"-", "--"}:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def is_positive_number(value):
    number = parse_number(value)
    return isinstance(number, (int, float)) and number > 0


def parse_percent_to_decimal(value):
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return value if value <= 1 else value / 100
    text = str(value).strip()
    if not text or text in {"-", "--"}:
        return None
    if text.endswith("%"):
        try:
            return round(float(text[:-1]) / 100, 4)
        except ValueError:
            return None
    try:
        number = float(text)
        return number if number <= 1 else number / 100
    except ValueError:
        return None


def find_header_row(ws, required_headers):
    required = {normalize_header(h) for h in required_headers}
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 20)):
        headers = {normalize_header(cell.value) for cell in row}
        if required.issubset(headers):
            return row[0].row
    raise ValueError(f"找不到包含表头 {required_headers} 的行：{ws.title}")


def build_header_map(ws, header_row, parent_rows=None):
    header_map = {}
    for cell in ws[header_row]:
        name = normalize_header(cell.value)
        if not name:
            continue
        header_map.setdefault(name, []).append(cell.column)
        if parent_rows:
            parts = []
            for row_number in parent_rows:
                parent = normalize_header(ws.cell(row=row_number, column=cell.column).value)
                if parent:
                    parts.append(parent)
            parts.append(name)
            header_map.setdefault("|".join(parts), []).append(cell.column)
    return header_map


def first_col(header_map, *names):
    for name in names:
        key = normalize_header(name)
        if key in header_map and header_map[key]:
            return header_map[key][0]
    return None


def get_by_header(row_values, header_map, *names):
    column = first_col(header_map, *names)
    if not column or column - 1 >= len(row_values):
        return None
    return row_values[column - 1]


def make_unique_key(data, header_map):
    blogger_id = normalize_key(get_by_header(data, header_map, "博主ID"))
    account_id = normalize_key(get_by_header(data, header_map, "小红书号", "账号ID"))
    home_url = normalize_key(get_by_header(data, header_map, "小红书主页", "主页链接"))
    return blogger_id or account_id or home_url


def read_pgy_rows(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    header_row = find_header_row(ws, ["博主ID", "小红书号", "昵称"])
    header_map = build_header_map(ws, header_row, parent_rows=[1, 2])
    rows = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not any(cell is not None and str(cell).strip() for cell in row):
            continue
        key = make_unique_key(row, header_map)
        if key:
            rows.append({"key": key, "values": row})
    return rows, header_map


def read_ratecard_manual_rows(ws, header_row, header_map):
    manual_by_key = {}
    order_by_key = {}
    raw_by_key = {}
    row_number = 0
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not any(cell is not None and str(cell).strip() for cell in row):
            continue
        row_number += 1
        key = make_unique_key(row, header_map)
        if not key:
            continue
        manual_by_key[key] = {field: get_by_header(row, header_map, field) for field in MANUAL_FIELDS}
        order_by_key[key] = row_number
        raw_by_key[key] = row
    return manual_by_key, order_by_key, raw_by_key


def pgy_metric(pgy_values, pgy_headers, metric_name):
    cooperation = get_by_header(
        pgy_values,
        pgy_headers,
        f"数据概览|笔记数据-按规模-合作笔记|{metric_name}",
    )
    if is_positive_number(cooperation):
        return parse_number(cooperation)
    daily = get_by_header(
        pgy_values,
        pgy_headers,
        f"数据概览|笔记数据-按规模-日常笔记|{metric_name}",
    )
    return parse_number(daily)


def rounded_divide(numerator, denominator, multiplier=1):
    numerator = parse_number(numerator)
    denominator = parse_number(denominator)
    if not isinstance(numerator, (int, float)) or not isinstance(denominator, (int, float)):
        return None
    if denominator <= 0:
        return None
    return round(numerator / denominator * multiplier, 2)


def pgy_price(pgy_values, pgy_headers, kind):
    if kind == "图文":
        return parse_number(get_by_header(pgy_values, pgy_headers, "图文笔记一口价"))
    if kind == "视频":
        return parse_number(get_by_header(pgy_values, pgy_headers, "视频笔记一口价"))
    return None


def pgy_org(pgy_values, pgy_headers):
    return normalize_key(get_by_header(pgy_values, pgy_headers, "所属机构"))


def is_allowed_org(org):
    return org in {"", "漫洵文化"}


def append_note(existing, note):
    if not note:
        return existing
    if existing is None or str(existing).strip() == "":
        return note
    if note in str(existing):
        return existing
    return f"{existing}；{note}"


def value_for_output(header, pgy_row, pgy_headers, manual_values, sequence, extra_note=None):
    header = normalize_header(header)
    pgy_values = pgy_row["values"]
    direct_map = {
        "昵称": ("昵称",),
        "账号类型": ("内容类型",),
        "账号ID": ("小红书号",),
        "主页链接": ("小红书主页",),
        "蒲公英链接": ("蒲公英主页",),
        "粉丝量/w": ("粉丝数（万）",),
        "地区": ("地理位置",),
        "内容标签": ("内容类型",),
    }

    if header == "序号":
        return sequence
    if header == ">18占比":
        under_18 = parse_percent_to_decimal(get_by_header(pgy_values, pgy_headers, "年龄分布-<18"))
        return None if under_18 is None else round(1 - under_18, 4)
    if header == "男性粉丝":
        return parse_percent_to_decimal(get_by_header(pgy_values, pgy_headers, "性别分布-男粉占比"))
    if header == "女性粉丝":
        return parse_percent_to_decimal(get_by_header(pgy_values, pgy_headers, "性别分布-女粉占比"))
    if header == "曝光中位数":
        return pgy_metric(pgy_values, pgy_headers, "曝光中位数")
    if header == "阅读中位数":
        return pgy_metric(pgy_values, pgy_headers, "阅读中位数")
    if header == "互动中位数":
        return pgy_metric(pgy_values, pgy_headers, "互动中位数")
    if header == "图文笔记一口价（含税不含平台费）":
        return pgy_price(pgy_values, pgy_headers, "图文")
    if header == "视频笔记一口价（含税不含平台费）":
        return pgy_price(pgy_values, pgy_headers, "视频")
    if header == "图文预估cpm":
        return rounded_divide(pgy_price(pgy_values, pgy_headers, "图文"), pgy_metric(pgy_values, pgy_headers, "曝光中位数"), 10000)
    if header == "视频预估cpm":
        return rounded_divide(pgy_price(pgy_values, pgy_headers, "视频"), pgy_metric(pgy_values, pgy_headers, "曝光中位数"), 10000)
    if header == "图文预估cpe":
        return rounded_divide(pgy_price(pgy_values, pgy_headers, "图文"), pgy_metric(pgy_values, pgy_headers, "互动中位数"))
    if header == "视频预估cpe":
        return rounded_divide(pgy_price(pgy_values, pgy_headers, "视频"), pgy_metric(pgy_values, pgy_headers, "互动中位数"))
    if header == "机构":
        return pgy_org(pgy_values, pgy_headers) or manual_values.get(header)
    if header == "备注":
        return append_note(manual_values.get(header), extra_note)
    if header in MANUAL_FIELDS:
        return manual_values.get(header)
    if header in direct_map:
        value = get_by_header(pgy_values, pgy_headers, *direct_map[header])
        return parse_number(value) if header == "粉丝量/w" else value
    return None


def copy_row_style(ws, source_row, target_row, max_col):
    for col in range(1, max_col + 1):
        src = ws.cell(row=source_row, column=col)
        dst = ws.cell(row=target_row, column=col)
        if src.has_style:
            dst._style = copy.copy(src._style)
        dst.number_format = src.number_format
        dst.alignment = copy.copy(src.alignment)
        dst.font = copy.copy(src.font)
        dst.fill = copy.copy(src.fill)
        dst.border = copy.copy(src.border)


def clear_and_write_rows(ws, header_row, headers, output_rows):
    template_row = header_row + 1
    max_col = len(headers)
    if ws.max_row > template_row:
        ws.delete_rows(template_row + 1, ws.max_row - template_row)
    for col in range(1, max_col + 1):
        ws.cell(row=template_row, column=col).value = None
    if not output_rows:
        return
    if len(output_rows) > 1:
        ws.insert_rows(template_row + 1, len(output_rows) - 1)
    for row_offset, row_values in enumerate(output_rows):
        target_row = template_row + row_offset
        if target_row != template_row:
            copy_row_style(ws, template_row, target_row, max_col)
        for col, value in enumerate(row_values, start=1):
            ws.cell(row=target_row, column=col).value = value


def recreate_sheet(wb, title):
    if title in wb.sheetnames:
        del wb[title]
    return wb.create_sheet(title)


def write_inactive_sheet(wb, headers, inactive_rows):
    ws = recreate_sheet(wb, INACTIVE_SHEET)
    ws.append(["说明", "旧刊例中存在，但本次蒲公英导出未出现；默认不进入本期刊例主表。"])
    ws.append(headers + ["下架原因", "处理时间"])
    now = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    for row in inactive_rows:
        ws.append(list(row[: len(headers)]) + ["本次蒲公英导出未出现", now])


def write_other_org_sheet(wb, headers, rows):
    ws = recreate_sheet(wb, OTHER_ORG_SHEET)
    ws.append(["说明", "蒲公英导出中所属机构不是空、也不是漫洵文化；默认不进入本期刊例主表。"])
    ws.append(headers + ["不展示原因", "处理时间"])
    now = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    for row in rows:
        ws.append(row + ["所属机构非空且不是漫洵文化", now])


def write_duplicate_sheet(wb, headers, rows):
    ws = recreate_sheet(wb, DUPLICATE_SHEET)
    ws.append(["说明", "蒲公英导出中出现重复达人；主刊例默认保留第一条，重复明细在这里核对。"])
    ws.append(["重复键", "重复序号"] + headers + ["处理备注"])
    for duplicate_key, duplicate_index, row in rows:
        ws.append([duplicate_key, duplicate_index] + row + ["重复达人，主刊例默认保留第一条"])


def write_log_sheet(wb, stats, args):
    ws = recreate_sheet(wb, LOG_SHEET)
    rows = [
        ("更新时间", dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("蒲公英文件", str(args.pgy)),
        ("旧刊例文件", str(args.ratecard)),
        ("输出文件", str(args.output)),
        ("本次蒲公英达人数量", stats["pgy_count"]),
        ("生成刊例达人数量", stats["active_count"]),
        ("新增达人数量", stats["new_count"]),
        ("保留旧人工字段达人数量", stats["kept_manual_count"]),
        ("下架留存达人数量", stats["inactive_count"]),
        ("其他机构不展示数量", stats["other_org_count"]),
        ("重复达人记录数量", stats["duplicate_count"]),
    ]
    for row in rows:
        ws.append(row)


def generate(args):
    pgy_rows, pgy_headers = read_pgy_rows(args.pgy)
    wb = openpyxl.load_workbook(args.ratecard)
    if TARGET_SHEET not in wb.sheetnames:
        raise ValueError(f"旧刊例文件里找不到工作表：{TARGET_SHEET}")
    ws = wb[TARGET_SHEET]
    header_row = find_header_row(ws, ["序号", "昵称", "账号ID"])
    ratecard_headers = [cell.value for cell in ws[header_row] if cell.column <= ws.max_column]
    ratecard_header_map = build_header_map(ws, header_row)
    manual_by_key, order_by_key, raw_by_key = read_ratecard_manual_rows(ws, header_row, ratecard_header_map)

    pgy_by_key = {}
    pgy_position = {}
    duplicate_rows = []
    duplicate_keys = set()
    duplicate_counts = {}
    for position, row in enumerate(pgy_rows):
        key = row["key"]
        duplicate_counts[key] = duplicate_counts.get(key, 0) + 1
        if key in pgy_by_key:
            duplicate_keys.add(key)
            duplicate_rows.append((key, duplicate_counts[key], row))
            continue
        pgy_by_key[key] = row
        pgy_position[key] = position

    inactive_keys = [key for key in raw_by_key if key not in set(pgy_by_key)]
    candidate_rows = sorted(
        pgy_by_key.values(),
        key=lambda row: (order_by_key.get(row["key"], 10**9), pgy_position[row["key"]]),
    )

    output_rows = []
    other_org_rows = []
    duplicate_output_rows = []
    kept_manual_count = 0
    new_count = 0
    for pgy_row in candidate_rows:
        org = pgy_org(pgy_row["values"], pgy_headers)
        manual_values = manual_by_key.get(pgy_row["key"], {})
        extra_note = "蒲公英导入重复，已保留第一条" if pgy_row["key"] in duplicate_keys else None
        if not is_allowed_org(org):
            other_org_rows.append([
                value_for_output(header, pgy_row, pgy_headers, manual_values, len(other_org_rows) + 1, extra_note)
                for header in ratecard_headers
            ])
            continue
        if manual_values:
            kept_manual_count += 1
        else:
            new_count += 1
        output_rows.append([
            value_for_output(header, pgy_row, pgy_headers, manual_values, len(output_rows) + 1, extra_note)
            for header in ratecard_headers
        ])

    for duplicate_key, duplicate_index, pgy_row in duplicate_rows:
        manual_values = manual_by_key.get(duplicate_key, {})
        duplicate_output_rows.append((
            duplicate_key,
            duplicate_index,
            [
                value_for_output(header, pgy_row, pgy_headers, manual_values, duplicate_index, "蒲公英导入重复")
                for header in ratecard_headers
            ],
        ))

    clear_and_write_rows(ws, header_row, ratecard_headers, output_rows)
    write_inactive_sheet(wb, ratecard_headers, [raw_by_key[key] for key in inactive_keys])
    write_other_org_sheet(wb, ratecard_headers, other_org_rows)
    write_duplicate_sheet(wb, ratecard_headers, duplicate_output_rows)
    stats = {
        "pgy_count": len(pgy_rows),
        "active_count": len(output_rows),
        "new_count": new_count,
        "kept_manual_count": kept_manual_count,
        "inactive_count": len(inactive_keys),
        "other_org_count": len(other_org_rows),
        "duplicate_count": len(duplicate_rows),
    }
    write_log_sheet(wb, stats, args)
    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return stats


def main():
    parser = argparse.ArgumentParser(description="根据蒲公英导出表自动生成小红书刊例。")
    parser.add_argument("--pgy", required=True, type=Path, help="蒲公英导出的 Excel 文件路径")
    parser.add_argument("--ratecard", required=True, type=Path, help="旧刊例 Excel 文件路径")
    parser.add_argument("--output", required=True, type=Path, help="输出的新刊例 Excel 文件路径")
    args = parser.parse_args()
    stats = generate(args)
    print("生成完成")
    for key, value in stats.items():
        print(f"{key}: {value}")


if __name__ == "__main__":
    main()
