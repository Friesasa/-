from __future__ import annotations

import shutil
import uuid
import os
from pathlib import Path
from typing import Any

import openpyxl
from fastapi import UploadFile
from openpyxl.utils import get_column_letter

from generate_ratecard import (
    build_header_map,
    find_header_row,
    get_by_header,
    make_unique_key,
    normalize_key,
    parse_number,
    parse_percent_to_decimal,
    pgy_metric,
    pgy_org,
    pgy_price,
    read_pgy_rows,
    rounded_divide,
)

from .field_mapping import empty, match_standard_field, standard_label


BASE_DIR = Path(__file__).resolve().parents[1]
STORAGE_DIR = Path(os.getenv("STORAGE_DIR", BASE_DIR))
UPLOAD_DIR = Path(os.getenv("UPLOAD_DIR", STORAGE_DIR / "uploads"))
EXPORT_DIR = Path(os.getenv("EXPORT_DIR", STORAGE_DIR / "exports"))
INQUIRY_MATCH_ONLY_FIELDS = {"nickname"}
MISSING_TEXT_VALUES = {"missing", "未提及", "无", "none", "null", "n/a"}


def ensure_storage() -> None:
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)


async def save_upload(file: UploadFile, subdir: str, owner_id: int | None = None) -> Path:
    ensure_storage()
    suffix = Path(file.filename or "").suffix or ".xlsx"
    target_dir = UPLOAD_DIR / subdir
    if owner_id is not None:
        target_dir = target_dir / str(owner_id)
    target_dir.mkdir(parents=True, exist_ok=True)
    target = target_dir / f"{uuid.uuid4().hex}{suffix}"
    with target.open("wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
    return target


def import_ratecard(path: Path) -> dict[str, Any]:
    try:
        rows, headers = read_pgy_rows(path)
        talents = [_talent_from_pgy(row, headers) for row in rows]
        return {
            "sheet_name": openpyxl.load_workbook(path, read_only=True).active.title,
            "header_row": 0,
            "talents": [talent for talent in talents if talent["unique_key"]],
        }
    except Exception:
        return _import_generic_ratecard(path)


def _talent_from_pgy(row: dict[str, Any], headers: dict[str, list[int]]) -> dict[str, Any]:
    values = row["values"]
    data = {
        "nickname": get_by_header(values, headers, "昵称"),
        "account_type": get_by_header(values, headers, "内容类型"),
        "account_id": normalize_key(get_by_header(values, headers, "小红书号", "账号ID")),
        "blogger_id": normalize_key(get_by_header(values, headers, "博主ID")),
        "homepage": get_by_header(values, headers, "小红书主页", "主页链接"),
        "pgy_url": get_by_header(values, headers, "蒲公英主页", "蒲公英链接"),
        "followers_w": parse_number(get_by_header(values, headers, "粉丝数（万）")),
        "exposure_median": pgy_metric(values, headers, "曝光中位数"),
        "read_median": pgy_metric(values, headers, "阅读中位数"),
        "engagement_median": pgy_metric(values, headers, "互动中位数"),
        "image_price": pgy_price(values, headers, "图文"),
        "video_price": pgy_price(values, headers, "视频"),
        "image_cpm": rounded_divide(pgy_price(values, headers, "图文"), pgy_metric(values, headers, "曝光中位数"), 10000),
        "video_cpm": rounded_divide(pgy_price(values, headers, "视频"), pgy_metric(values, headers, "曝光中位数"), 10000),
        "image_cpe": rounded_divide(pgy_price(values, headers, "图文"), pgy_metric(values, headers, "互动中位数")),
        "video_cpe": rounded_divide(pgy_price(values, headers, "视频"), pgy_metric(values, headers, "互动中位数")),
        "female_ratio": parse_percent_to_decimal(get_by_header(values, headers, "性别分布-女粉占比")),
        "male_ratio": parse_percent_to_decimal(get_by_header(values, headers, "性别分布-男粉占比")),
        "region": get_by_header(values, headers, "地理位置"),
        "tags": get_by_header(values, headers, "内容类型"),
        "organization": pgy_org(values, headers),
    }
    under_18 = parse_percent_to_decimal(get_by_header(values, headers, "年龄分布-<18"))
    if under_18 is not None:
        data["adult_ratio"] = round(1 - under_18, 4)
    return {
        "unique_key": row["key"],
        "nickname": data.get("nickname"),
        "account_id": data.get("account_id"),
        "blogger_id": data.get("blogger_id"),
        "homepage": data.get("homepage"),
        "data": data,
    }


def _import_generic_ratecard(path: Path) -> dict[str, Any]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    best_ws = None
    best_header_row = 1
    for ws in wb.worksheets:
        try:
            best_header_row = find_header_row(ws, ["昵称"])
            best_ws = ws
            break
        except ValueError:
            continue
    if best_ws is None:
        best_ws = wb.active
        best_header_row = _detect_header_row(best_ws)
    headers = build_header_map(best_ws, best_header_row)
    talents = []
    for row in best_ws.iter_rows(min_row=best_header_row + 1, values_only=True):
        if not any(not empty(cell) for cell in row):
            continue
        key = make_unique_key(row, headers)
        data = _standard_data_from_headers(row, headers)
        if not key:
            key = normalize_key(data.get("account_id") or data.get("blogger_id") or data.get("homepage") or data.get("nickname"))
        if key:
            talents.append(
                {
                    "unique_key": key,
                    "nickname": data.get("nickname"),
                    "account_id": data.get("account_id"),
                    "blogger_id": data.get("blogger_id"),
                    "homepage": data.get("homepage"),
                    "data": data,
                }
            )
    return {"sheet_name": best_ws.title, "header_row": best_header_row, "talents": talents}


def _standard_data_from_headers(row: tuple[Any, ...], header_map: dict[str, list[int]]) -> dict[str, Any]:
    data: dict[str, Any] = {}
    for header, columns in header_map.items():
        field, score = match_standard_field(header)
        if not field or score < 0.58:
            continue
        value = row[columns[0] - 1] if columns[0] - 1 < len(row) else None
        if empty(value):
            continue
        data[field] = value
    for field in {"followers_w", "image_price", "video_price", "exposure_median", "read_median", "engagement_median"}:
        if field in data:
            data[field] = parse_number(data[field])
    return data


def create_brand_task(path: Path, brand_name: str | None = None) -> dict[str, Any]:
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    header_row = _detect_header_row(ws)
    columns = []
    for cell in ws[header_row]:
        header = cell.value
        if empty(header):
            continue
        standard_field, confidence = match_standard_field(header)
        columns.append(
            {
                "key": f"c{cell.column}",
                "index": cell.column,
                "letter": get_column_letter(cell.column),
                "header": str(header),
                "standard_field": standard_field,
                "standard_label": standard_label(standard_field),
                "confidence": confidence,
            }
        )
    rows = []
    for row in ws.iter_rows(min_row=header_row + 1, max_col=max((col["index"] for col in columns), default=ws.max_column)):
        if not any(not empty(cell.value) for cell in row):
            continue
        rows.append(
            {
                "excel_row": row[0].row,
                "values": {column["key"]: ws.cell(row=row[0].row, column=column["index"]).value for column in columns},
                "source": {},
            }
        )
    if not rows:
        rows.append({"excel_row": header_row + 1, "values": {column["key"]: "" for column in columns}, "source": {}})
    return {
        "brand_name": brand_name,
        "filename": path.name,
        "template_path": str(path),
        "sheet_name": ws.title,
        "header_row": header_row,
        "columns": columns,
        "rows": rows,
        "missing": compute_missing(columns, rows),
    }


def _detect_header_row(ws: Any) -> int:
    best_row = 1
    best_score = -1
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 20)):
        values = [cell.value for cell in row if not empty(cell.value)]
        if not values:
            continue
        mapped = sum(1 for value in values if match_standard_field(value)[0])
        score = len(values) + mapped * 2
        if score > best_score:
            best_score = score
            best_row = row[0].row
    return best_row


def fill_row_from_talent(task: dict[str, Any], row_index: int, talent: dict[str, Any]) -> dict[str, Any]:
    rows = task["rows"]
    if row_index < 0 or row_index >= len(rows):
        raise IndexError("row_index out of range")
    data = talent["data"]
    row = rows[row_index]
    for column in task["columns"]:
        field = effective_standard_field(column)
        if field and field in data and not empty(data[field]):
            row["values"][column["key"]] = data[field]
            row.setdefault("source", {})[column["key"]] = "刊例"
    task["missing"] = compute_missing(task["columns"], rows)
    return task


def apply_parsed_inquiry(task: dict[str, Any], row_index: int, parsed: dict[str, Any]) -> dict[str, Any]:
    rows = task["rows"]
    if row_index < 0 or row_index >= len(rows):
        raise IndexError("row_index out of range")
    fields = parsed.get("fields", {})
    row = rows[row_index]
    for column in task["columns"]:
        field = effective_standard_field(column)
        if field in INQUIRY_MATCH_ONLY_FIELDS:
            continue
        value = _cell_value_from_parsed_field(fields.get(field))
        if field in fields and not empty(value):
            row["values"][column["key"]] = value
            row.setdefault("source", {})[column["key"]] = "二询"
    row["inquiry_parse"] = parsed
    task["missing"] = compute_missing(task["columns"], rows)
    return task


def _cell_value_from_parsed_field(value: Any) -> Any:
    if isinstance(value, dict):
        value = value.get("value")
    if isinstance(value, str) and value.strip().lower() in MISSING_TEXT_VALUES:
        return None
    return value


def resolve_inquiry_row_index(
    task: dict[str, Any],
    parsed: dict[str, Any],
    match_query: str | None = None,
    confirmed_row_index: int | None = None,
    selected_row_index: int | None = None,
) -> tuple[int | None, dict[str, Any]]:
    fields = parsed.get("fields", {})
    nickname = fields.get("nickname")
    match_info: dict[str, Any] = {
        "parsed_has_nickname": bool(nickname),
        "manual_query_provided": bool(match_query and match_query.strip()),
        "confirmed_row_provided": confirmed_row_index is not None,
        "matched_by": None,
        "matched_row_index": None,
    }

    parsed_match = _find_unique_row_match(task, nickname, {"nickname"})
    if _is_unique_row_match(parsed_match):
        match_info.update(parsed_match)
        return parsed_match["matched_row_index"], match_info
    selected_parsed_match = _select_from_ambiguous_match(parsed_match, selected_row_index, "selected_parsed_identity")
    if selected_parsed_match:
        match_info.update(selected_parsed_match)
        return selected_parsed_match["matched_row_index"], match_info

    manual_match = _find_unique_row_match(task, match_query, {"nickname", "account_id", "blogger_id", "homepage"})
    if _is_unique_row_match(manual_match):
        manual_match["matched_by"] = f"manual_{manual_match['matched_by']}"
        match_info.update(manual_match)
        return manual_match["matched_row_index"], match_info
    selected_manual_match = _select_from_ambiguous_match(manual_match, selected_row_index, "selected_manual_identity")
    if selected_manual_match:
        match_info.update(selected_manual_match)
        return selected_manual_match["matched_row_index"], match_info

    if confirmed_row_index is not None and 0 <= confirmed_row_index < len(task["rows"]):
        match_info["matched_by"] = "confirmed_selected_row"
        match_info["matched_row_index"] = confirmed_row_index
        return confirmed_row_index, match_info

    match_info.update(manual_match if manual_match["matched_row_index"] is not None else parsed_match)
    return None, match_info


def _is_unique_row_match(match_info: dict[str, Any]) -> bool:
    return isinstance(match_info.get("matched_row_index"), int) and not str(match_info.get("matched_by", "")).startswith("ambiguous")


def _select_from_ambiguous_match(match_info: dict[str, Any], selected_row_index: int | None, matched_by: str) -> dict[str, Any] | None:
    matched_rows = match_info.get("matched_row_index")
    if selected_row_index is None or not isinstance(matched_rows, list):
        return None
    if selected_row_index not in matched_rows:
        return None
    selected_match = dict(match_info)
    selected_match["matched_by"] = matched_by
    selected_match["matched_row_index"] = selected_row_index
    selected_match["ambiguous_candidates"] = matched_rows
    return selected_match


def _find_unique_row_match(task: dict[str, Any], query: Any, fields: set[str]) -> dict[str, Any]:
    rows = task["rows"]
    match_columns = [
        column
        for column in task.get("columns", [])
        if effective_standard_field(column) in fields
    ]
    match_info: dict[str, Any] = {
        "match_column_keys": [column["key"] for column in match_columns],
        "matched_by": None,
        "matched_row_index": None,
    }
    normalized_query = _normalize_match_text(query)
    if not normalized_query or not match_columns:
        return match_info

    exact_matches: list[int] = []
    partial_matches: list[int] = []
    for index, row in enumerate(rows):
        row_values = [
            _normalize_match_text(row.get("values", {}).get(column["key"]))
            for column in match_columns
        ]
        if normalized_query in row_values:
            exact_matches.append(index)
        elif any(normalized_query in row_value or row_value in normalized_query for row_value in row_values if row_value):
            partial_matches.append(index)

    if len(exact_matches) == 1:
        match_info["matched_by"] = "exact_identity"
        match_info["matched_row_index"] = exact_matches[0]
    elif len(exact_matches) > 1:
        match_info["matched_by"] = "ambiguous_exact_identity"
        match_info["matched_row_index"] = exact_matches
    elif len(partial_matches) == 1:
        match_info["matched_by"] = "partial_identity"
        match_info["matched_row_index"] = partial_matches[0]
    elif len(partial_matches) > 1:
        match_info["matched_by"] = "ambiguous_partial_identity"
        match_info["matched_row_index"] = partial_matches
    return match_info


def _normalize_match_text(value: Any) -> str:
    if value is None:
        return ""
    return "".join(str(value).split()).lower()


def effective_standard_field(column: dict[str, Any]) -> str | None:
    header = column.get("header")
    field, score = match_standard_field(header)
    if field and score >= 0.58:
        return field
    if not empty(header):
        return None
    return column.get("standard_field")


def compute_missing(columns: list[dict[str, Any]], rows: list[dict[str, Any]]) -> dict[str, list[str]]:
    missing: dict[str, list[str]] = {}
    required_columns = [column for column in columns if effective_standard_field(column)]
    for index, row in enumerate(rows):
        empty_headers = [
            column["header"]
            for column in required_columns
            if empty(row.get("values", {}).get(column["key"]))
        ]
        if empty_headers:
            missing[str(index)] = empty_headers
    return missing


def export_task(task: dict[str, Any]) -> Path:
    ensure_storage()
    template = Path(task["template_path"])
    wb = openpyxl.load_workbook(template)
    ws = wb[task["sheet_name"]]
    for row in task["rows"]:
        excel_row = int(row["excel_row"])
        for column in task["columns"]:
            ws.cell(row=excel_row, column=column["index"]).value = row["values"].get(column["key"])
    target_dir = EXPORT_DIR / str(task.get("user_id") or "shared")
    target_dir.mkdir(parents=True, exist_ok=True)
    output = target_dir / f"brand_task_{task['id']}_{uuid.uuid4().hex[:8]}.xlsx"
    wb.save(output)
    return output

