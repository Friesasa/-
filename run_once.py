from __future__ import annotations

import datetime as dt
from pathlib import Path
from types import SimpleNamespace

import openpyxl

from generate_ratecard import generate


BASE_DIR = Path(__file__).resolve().parent
INPUT_DIR = BASE_DIR / "input"
OUTPUT_DIR = BASE_DIR / "output"


def sheet_contains_headers(ws, headers: set[str]) -> bool:
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 10), values_only=True):
        values = {str(value).replace("\n", "").replace(" ", "").strip() for value in row if value is not None}
        if headers.issubset(values):
            return True
    return False


def detect_file(path: Path) -> str | None:
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return None

    if wb.worksheets and sheet_contains_headers(wb.worksheets[0], {"博主ID", "小红书号", "昵称"}):
        return "pgy"

    if "5月刊例" in wb.sheetnames:
        ws = wb["5月刊例"]
        if sheet_contains_headers(ws, {"序号", "昵称", "账号ID"}):
            return "ratecard"

    return None


def main() -> None:
    INPUT_DIR.mkdir(exist_ok=True)
    OUTPUT_DIR.mkdir(exist_ok=True)

    excel_files = [
        path
        for path in INPUT_DIR.iterdir()
        if path.suffix.lower() in {".xlsx", ".xlsm"} and not path.name.startswith("~$")
    ]

    pgy_files = []
    ratecard_files = []
    unknown_files = []

    for path in excel_files:
        file_type = detect_file(path)
        if file_type == "pgy":
            pgy_files.append(path)
        elif file_type == "ratecard":
            ratecard_files.append(path)
        else:
            unknown_files.append(path)

    if len(pgy_files) != 1 or len(ratecard_files) != 1:
        print("没有成功识别两个文件。")
        print("")
        print("请确认 input 文件夹里只有：")
        print("1. 一份蒲公英导出的 Excel")
        print("2. 一份旧刊例 Excel")
        print("")
        print(f"识别到蒲公英文件：{len(pgy_files)} 个")
        for path in pgy_files:
            print(f"  - {path.name}")
        print(f"识别到旧刊例文件：{len(ratecard_files)} 个")
        for path in ratecard_files:
            print(f"  - {path.name}")
        if unknown_files:
            print(f"无法识别文件：{len(unknown_files)} 个")
            for path in unknown_files:
                print(f"  - {path.name}")
        raise SystemExit(1)

    timestamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = OUTPUT_DIR / f"小红书刊例_自动生成_{timestamp}.xlsx"

    args = SimpleNamespace(pgy=pgy_files[0], ratecard=ratecard_files[0], output=output_path)
    stats = generate(args)

    print("生成完成")
    print(f"蒲公英文件：{pgy_files[0].name}")
    print(f"旧刊例文件：{ratecard_files[0].name}")
    print(f"输出文件：{output_path}")
    print("")
    for key, value in stats.items():
        print(f"{key}: {value}")


if __name__ == "__main__":
    main()
